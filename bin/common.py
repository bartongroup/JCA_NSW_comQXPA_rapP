'''
Functions reused in multiple scripts live here...
'''

import sys
import re
from json import load,loads
import gzip
from pathlib import Path
import subprocess

from lxml import etree
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from Bio.SeqRecord import SeqRecord
from Bio.Seq import Seq
from Bio import SeqIO

import requests
from requests.adapters import HTTPAdapter, Retry

def make_request(uri):

    """
    Makes HTTP request and returns result body

    Required params:
        uri(str): URI for request

    Returns:
        text of response
    """

    s = requests.Session()

    retries = Retry(total=10,
              backoff_factor=2,
              status_forcelist=[ 500, 502, 503, 504 ],
              raise_on_status=True)

    s.mount('https://', HTTPAdapter(max_retries=retries))

    try:
        r = s.get(uri, timeout=30)
    except requests.exceptions.RequestException as errex:
        print(f"Exception: {uri} - {errex}")

    return r.text


def get_taxa_name(ncbi_taxid):

    """
    Looks up taxa via ENA browser API to retrieve scientific name

    Required parameters:
        ncbi_taxid(str): Taxonomy identifier of species

    Returns:
        species_name(str): parsed scientific name
    """

    uri = f"https://www.ebi.ac.uk/ena/browser/api/xml/taxon:{ncbi_taxid}"
    result = make_request(uri)

    # lxml does not support embedded encodings...
    result = result.replace(' encoding="UTF-8"', '')
    root = etree.fromstring(result)

    error = root.xpath("/ErrorDetails/status")
    if error:
        print("Error obtaining taxonomy lookup")
        print(root)
        sys.exit(1)

    taxon = root.xpath("/TAXON_SET/taxon")[0]
    species_name = taxon.get('scientificName')

    return species_name

def clean_description(description):

    """
    tidies up variously formatted descriptions into something reasonably consistent

    Required params:
        description(str): record description
    
    Returns:
        description(str): cleaned description
    """

    description = re.sub(r'[, ]*complete genome[.]*', '', description)
    description = re.sub(r'[, ]*genome assembly[,:_ A-Za-z0-9]*', '', description)
    description = re.sub(r'chromosome[ ,]*', '', description)
    description = re.sub('whole genome shotgun sequence.','', description)
    description = re.sub(r'NODE[A-Za-z0-9_,]*', '', description)
    description = re.sub(r'[ ,]*$','', description)

    return description

def clean_strain(strain):

    """
    tidies up variously formatted strain names into something reasonably consistent

    Required params:
        strain(str): strain name
    
    Returns:
        strain(str): cleaned strain name
    """

    strain = re.sub(r'[, ]*(Mutant )?Bacillus subtilis (strain|isolate) [.]*', '', strain)
    strain = re.sub(r'[, ]*(Mutant )?Bacillus subtilis [.]*', '', strain)
    strain = re.sub(r' (strain|isolate)', '', strain)

    return strain

def get_busco_completeness(busco_lineage, accession):

    """
    Obtains busco completeness percentage for a provided genome accession

    Required params: 
        busco_lineage(str): BUSCO lineage used for classification
        accession(str): Genome accession

    Returns:
        completeness(float): BUSCO completeness percentage
    """

    busco_result = f'data/full/busco/{accession}/short_summary.specific.{busco_lineage}.{accession}.json'

    with open(busco_result, encoding='UTF-8') as fh:
        dat = load(fh)
        completeness = dat.get('results').get('Complete percentage')

    return completeness

def join_non_empty(vals):

    """
    Joins strings passed as list with semicolons for elements which are not empty,

    Required params: 
        vals(list): list to join

    Returns
        text(str): joined string, or single value
    """

    vals = list(filter(lambda x: x != "", vals))

    return "; ".join(vals)


def get_expt_accession(project, biosample):

    """
    Retrieves experiment accession from ENA for a given project and biosample

    Required params:
        project(str): PRJNA accession
        biosample(str): biosample accession

    Returns:
        expt_ids(str): comma-separated list of ENA run accession
    """

    project_json = Path(f"data/full/metadata/ena_runs/{project}.json")
    if project_json.is_file():
        with open(project_json, 'r', encoding='UTF-8') as fh:
            result = fh.read()
    else:
        uri = f"https://www.ebi.ac.uk/ena/portal/api/filereport?accession={project}&result=read_run&fields=sample_accession,experiment_accession,run_accession&format=json&limit=0"
        result = make_request(uri)
        with open(f"data/full/metadata/ena_runs/{project}.json", 'w', encoding='UTF-8') as fh:
            fh.write(result)

    data = loads(result)
    expt_accessions = []
    for run in data:
        if run['sample_accession'] == biosample:
            expt_accessions.append(run['experiment_accession'])

    return ','.join(expt_accessions)

def get_sequence_tech(run_ids):

    """
    Retrieves run data extracts sequencing platform 

    Required params:
        run_ids(str): comma-separated list of ENA run ids

    Returns:
        platorms: comma-separated list of sequencing platforms
    """

    platforms = []
    for run_id in run_ids.split(','):
        if run_id:
            run_file = Path(f'data/full/metadata/ena_expts/{run_id}.xml')
            if  run_file.exists():
                with open(run_file, 'r', encoding='UTF-8') as fh:
                    result = fh.read()
            else:
                uri = f"https://www.ebi.ac.uk/ena/browser/api/xml/{run_id}"
                result = make_request(uri)
                with open(run_file, 'w', encoding='UTF-8') as fh:
                    fh.write(result)

            if result:
                result = result.replace(' encoding="UTF-8"', '')
                root = etree.fromstring(result)

                # platform should be stored as <PLATFORM><$PLATFORMNAME>
                platform = root.xpath("EXPERIMENT/PLATFORM/*")
                if len(platform):
                    platforms.append(platform[0].tag)

    platforms = list(set(platforms))

    return ','.join(platforms)

def get_ena_tech(accession):
    """
    Parses sequencing platform from ENA entry

    Required arguments
        accession(str): ENA genome accession
    
    Returns:
        platform(str): Sequencing platform
    """
    with gzip.open(f'data/full/genomes/{accession}.embl.gz', 'rt') as fh:
        records = list(SeqIO.parse(fh, format='embl'))
        record = records[0]
        if 'comment' in record.annotations:
            comment = record.annotations['comment']

            result = re.search(r'Sequencing Technology *:: *([A-Za-z0-9; ]+)', comment)
            if result:
                return result.group(1)

    return None

def format_header_cell(ws, cell, label):

    """
    Formats a header cell

    Required Parameters: 
        ws: worksheet
        cell: cell to update
        label: text to set as cell value
    
    Returns:
        None
    """

    ws[cell].value = label
    ws[cell].font = Font(bold = True)
    ws[cell].alignment = Alignment(horizontal = 'center', vertical = 'center')


def combine_metadata(metadata, proteins, protein_coverage, classifications, busco_lineage, busco_threshold, output, accessions_file):

    """
    Update metadata sheet to add GTDBTK classifications and BUSCO completeness, 
    and add columns to indicate wheter genome should be retained and 
    reason[s] for exclusions

    Required params: 
        metadata(str): Path to metadata file created by 00_build_genome_dbs.py
        classifications(str): Path to GTDBTK classification output file
        busco_lineage(str): BUSCO lineage 
        busco_threshold(int): Proportion of complete BUSCOs required for inclusion
        output(str): Path to output file to create

    Returns:
        None
    """
    # TODO _ this should be somewhere else...but where?
    run_dir = Path('data/full/metadata/ena_runs')
    expt_dir = Path('data/full/metadata/ena_expts')
    run_dir.mkdir(exist_ok=True)
    expt_dir.mkdir(exist_ok=True)

    metadata = pd.read_csv(metadata, sep="\t")
    classifications = pd.read_csv(classifications, sep="\t")
    protein_df = pd.read_csv(proteins, sep="\t")
    coverage_df = pd.read_csv(protein_coverage, sep="\t")

    # Projects may be non-unique due to multiple assemblies being available for a project,
    # so make these distinct...
    metadata.drop_duplicates(subset='Accession', keep='first', inplace=True)

    # Sequencing platform may either be indicated in the SRA records, or in teh comments section of the ENA record
    # Need to check both of these and summarise...
    metadata['Experiment IDs'] = metadata.apply(lambda x: get_expt_accession(x['Study ID'], x['Biosample ID']), axis = 1)
    metadata['SRA_Platform'] = metadata.apply(lambda x: get_sequence_tech(x['Experiment IDs']), axis = 1)
    metadata['ENA_Platform'] = metadata.apply(lambda x: get_ena_tech(x['Accession']), axis = 1)
    metadata['Platform'] = metadata['ENA_Platform'].fillna(metadata['SRA_Platform'])

    metadata=metadata[['Accession', 'Study ID', 'Biosample ID', 'Experiment IDs', 'Title', 'Strain', 'Culture collection',
        'Host', 'Isolation source', 'Environmental context', 'Location', 'Collection date', 'Scaffolds', 'Platform']]

    # Add species column from GTDBTK outputs
    classifications['Species'] = classifications['classification'].map(lambda x: x.split(';')[-1].replace('s__',''))
    classifications = classifications[['user_genome','Species']]

    metadata = pd.merge(metadata, classifications, how='left', left_on='Accession', right_on='user_genome')
    metadata = metadata.drop('user_genome', axis = 1)

    # Add BUSCO completeness...
    metadata['BUSCO completeness'] = metadata['Accession'].map(lambda x: get_busco_completeness(busco_lineage, x))
    metadata.fillna(value = 'None', inplace = True)

    # Add 'Retain' column defining accessions to be retained, and column explaining rationale for this
    metadata['Retain'] = 1
    metadata['GTDBTK exclusion'] = ""
    metadata['BUSCO exclusion'] = ""
    metadata['mutant exclusion'] = ""

    metadata.loc[metadata['Species'] != "Bacillus subtilis", 'Retain'] = 0
    metadata.loc[metadata['Species'] != "Bacillus subtilis", 'GTDBTK exclusion'] = 'GTDBTK classification'
    metadata.loc[metadata['BUSCO completeness'] < busco_threshold, 'Retain'] = 0
    metadata.loc[metadata['BUSCO completeness'] < busco_threshold, 'BUSCO exclusion'] = 'Below BUSCO threshold'
    metadata.loc[metadata['Title'].str.contains('Mutant', case = False), 'mutant exclusion'] = 'Mutated isolate'
    metadata.loc[metadata['Title'].str.contains('Mutant', case = False), 'Retain'] = 0
    metadata.loc[metadata['Isolation source'].str.contains('Genome-engineer', case = False), 'mutant exclusion'] = 'Mutated isolate'
    metadata.loc[metadata['Isolation source'].str.contains('Genome-engineer', case = False), 'Retain'] = 0

    metadata["Exclusion criteria"] = metadata[["GTDBTK exclusion", "BUSCO exclusion", "mutant exclusion"]].apply(join_non_empty, axis=1)
    metadata.drop(['GTDBTK exclusion', 'BUSCO exclusion', 'mutant exclusion'], axis = 1, inplace = True)

    metadata = pd.merge(metadata, protein_df, left_on='Accession', right_on='accession', how='left')
    metadata = pd.merge(metadata, coverage_df, left_on='Accession', right_on='accession', how='left')
    metadata.drop(['accession_x','accession_y'], axis = 1, inplace = True)

    metadata.to_excel(output, index = False)

    accessions = metadata[metadata['Retain'] == 1]['Accession']
    accessions.to_csv(accessions_file, sep = "\t", index = False)

    # Reread excel file and tweak...
    wb = load_workbook(output)
    ws = wb.active

    ws.insert_rows(1)
    ws.merge_cells('S1:AC1')
    ws.merge_cells('AD1:AN1')
    ws.merge_cells('AO1:AY1')
    ws.merge_cells('AZ1:BJ1')
    ws.merge_cells('BK1:BU1')
    ws.merge_cells('BV1:CI1')
    ws.merge_cells('CJ1:CW1')
    ws.merge_cells('CX1:DK1')
    ws.merge_cells('DL1:DY1')
    ws.merge_cells('DZ1:EM1')

    format_header_cell(ws, 'S1', 'comA Protein Details')
    format_header_cell(ws, 'AD1', 'comP Protein Details')
    format_header_cell(ws, 'AO1', 'comX Protein Details')
    format_header_cell(ws, 'AZ1', 'comQ Protein Details')
    format_header_cell(ws, 'BK1', 'rapP Protein Details')

    format_header_cell(ws, 'BV1', 'comA Protein Blast results')
    format_header_cell(ws, 'CJ1', 'comP Protein Blast results')
    format_header_cell(ws, 'CX1', 'comX Protein Blast results')
    format_header_cell(ws, 'DL1', 'comQ Protein Blast results')
    format_header_cell(ws, 'DZ1', 'rapP Protein Blast results')

    wb.save(output)

def get_qualifier(name, feature):

    """
    Extracts a qualifier value from a feature for a given name

    Required params:
        name(str): qualifier name
        feature(Bio::SeqFeature): feature of interest
    """

    if feature.qualifiers.get(name):
        return feature.qualifiers.get(name)[0]

    return None


def extract_feature_details(feature):

    """
    Pulls required fields from feature object

    Required parameters: 
        feature(Bio::SeqFeature): feature of interest
    
    returns:
        feature_info(dict): Dictionary containing required fields
    """

    feature_info = {
        'location':  f"{feature.location.start}-{feature.location.end}"
    }

    translation = get_qualifier('translation', feature)
    feature_info['translation'] = translation
    if translation is not None:
        feature_info['cds_length'] = len(translation)

    feature_info['gene_id'] = get_qualifier('gene', feature)
    feature_info['product'] = get_qualifier('product', feature)

    feature_info['pseudogene'] = get_qualifier('pseudogene', feature)

    if feature.qualifiers.get('note') and feature_info['pseudogene']:
        feature_info['note'] = feature.qualifiers.get('note')[-1]
    else:
        feature_info['note'] = None

    feature_info['locus_tag'] = get_qualifier('locus_tag', feature)

    return feature_info

def extract_cds_details(cds_info, record, accession, gene_name):

    """
    Extracts details from CDS features, including creating sequence
    object from translation

    Required fields:
        cds_info(dict): Dict of details of CDS
        record(SeqRecord): Parent record for CDS feature
        accession(str): Genome accession

    Returns:
        cds_info(dict): Updated dict
        protein(Bio::SeqRecord): Sequence of protein translation
    """

    protein = None

    cds_info['record_id'] = record.id
    cds_info['genomic_context'] = 'chromosome'
    if 'plasmid' in record.description:
        cds_info['genomic_context'] = 'plasmid'
    gene_id = cds_info.get("gene_id")
    if gene_id is None:
        gene_id = gene_name
    gene_id = f'{accession}_{gene_id}'

    if cds_info.get('translation'):

        protein = SeqRecord(
            Seq(cds_info['translation']),
                id = gene_id,
                description = cds_info.get('product'),
                annotations={'molecule_type': "protein"}
        )

    return(cds_info, protein)

def get_target_sequences(genome):

    """
    Parses the genome record to extract required sequences

    required parameters:
        genome(libpath.Path): Path to record

    returns:
        accession(str): accession of record
        gene_info(dict): metadata dictionary keyed on gene name
        cds_seqs(dict): dict of Bio::Seq objects keyed on gene name
    """

    gene_info = {}
    cds_seqs = {}
    prot_seqs = {}

    accession = str(genome.stem).replace('.embl','')
    with open(genome, 'rt', encoding='UTF-8') as fh:

        for record in SeqIO.parse(fh, format = 'embl'):

        #comP and comQ are not reliably annotated, however comX is present in
        #every annotation so use this to identify the location of the operon

            for index, feature in enumerate(record.features):

                if feature.type == 'gene' and 'gene' in feature.qualifiers.keys():
                    if 'comX' in feature.qualifiers.get('gene')[0]:
                        comX_gene_index = index
                        cds_indices = {}

                        # The operon is generally on the -ve strand, but there are
                        # some cases where it is positive

                        # Since we have consistent annotation formats, we should
                        # be able to rely on the surrounding features being the
                        # comX CDS, and the comP and comQ genes and CDSs, which
                        # helps given some sequences are fairly divergent

                        if feature.location.strand== -1:

                            gene_info['strand']='-'
                            cds_indices['comX'] = comX_gene_index + 1
                            cds_indices['comP'] = comX_gene_index - 1
                            cds_indices['comQ'] = comX_gene_index + 3
                            cds_indices['comA'] = comX_gene_index - 3
                        else:

                            gene_info['strand']='+'
                            cds_indices['comX'] = comX_gene_index + 1
                            cds_indices['comP'] = comX_gene_index + 3
                            cds_indices['comQ'] = comX_gene_index - 1
                            cds_indices['comA'] = comX_gene_index + 5

                        for index, feature in cds_indices.items():

                            cds_info = extract_feature_details(record.features[feature])
                            # comP frameshifts can not only cause a truncation, but also a second product from the 3' region,
                            # which screws up comA positioning, as do transposase insertions within comP

                            # These checks work with all available genomes at the time of writing but changes may be required
                            # in the light of novel sequences being produced
                            if index == 'comA':
                                if not cds_info['product']:
                                    print(f"{record.id}: No product available" )
                                elif 'histidine kinase' in cds_info['product'] or 'hypothetical protein' in cds_info['product']:
                                    if gene_info['strand']=='-':
                                        cds_info = extract_feature_details(record.features[comX_gene_index-5])
                                    else:
                                        cds_info = extract_feature_details(record.features[comX_gene_index+7])
                                elif 'transposase' in cds_info['product']:
                                    if gene_info['strand']=='-':
                                        cds_info = extract_feature_details(record.features[comX_gene_index-7])
                                    else:
                                        cds_info = extract_feature_details(record.features[comX_gene_index+9])

                            if index == 'comQ':
                                if 'transposase' in cds_info['product']:
                                    if 'product' in cds_info and cds_info['product'] == "IS4 transposase":
                                        cds_info = extract_feature_details(record.features[comX_gene_index+7])
                                    elif gene_info['strand']=='-':
                                        cds_info = extract_feature_details(record.features[comX_gene_index+2])
                                    else:
                                        print(f"Unhandled trasnsposase found...{str(cds_info)}")

                            gene_info[index], prot_seqs[index] = extract_cds_details(cds_info, record, accession, index)

                # rapP does not seem to be annotated with a gene name by bakta, but does have
                # a product attached to CDS features
                elif feature.type == 'CDS' and 'product' in feature.qualifiers.keys() and \
                    feature.qualifiers.get('product')[0] == 'Rap phosphatase':

                    cds_info = extract_feature_details(feature)
                    gene_info['rapP'], prot_seqs['rapP'] = extract_cds_details(cds_info, record, accession, 'rapP')

    return(accession, gene_info, cds_seqs, prot_seqs)

def blast_index(fasta_dir, blast_dir, species, ncbi_taxid, index_type):
    """
    Creates a blast index for each genome fasta file contained within directory

    Required parameters:
        fasta_dir(Path): Location of fasta files
        blast_dir(Path): Location of blast database
        species(str): Species name
        ncbi_taxid(str): NCBI taxonomy ID
        index_type(str) : index type (nucl/prot)

    Returns:
        None
    """
    species_name=species.replace(' ','_')

    if index_type not in ['nucl','prot']:
        raise ValueError(f'Invalid index type provided ({index_type})')

    accessions = []

    db_stats = {'count': 0, 'length': 0}

    if index_type == 'nucl':
        title = f'TITLE {species} complete genomes'
        meta_db = f'{blast_dir}/{species_name}_complete_genomes.nal'
    else:
        title  =  f"TITLE {species} complete proteome"
        meta_db = f'{blast_dir}/{species_name}_proteins.pal'

    fasta_files = fasta_dir.glob('*.fasta')
    for fasta_file in fasta_files:
        with open(fasta_file, 'r', encoding='UTF-8') as fh:
            records = SeqIO.parse(fh, format = 'fasta')
            for record in records:
                db_stats['count'] += 1
                db_stats['length'] += len(record.seq)

        accession = fasta_file.name.replace('.fasta','')
        accessions.append(accession)

        cmd = ["makeblastdb",  "-in",  fasta_file, "-dbtype", index_type, "-out", blast_dir / Path(accession),
               "-taxid", str(ncbi_taxid), "-title", accession]

        try:
            subprocess.run(cmd, check = True, capture_output = False)
        except subprocess.CalledProcessError as e:
            print(f"Index failed: {accession} - {e}")

    accessions = ' '.join([f"{a}" for a in accessions])
    index = f'''\
{title}
DBLIST {accessions}
NSEQ {db_stats['count']}
LENGTH {db_stats['length']}
'''.strip()

    with open(meta_db, 'w', encoding='UTF-8') as fh:
        fh.writelines(index)

def copy_genome(accession):
    """
    Symlinks genome data for selected accession into 'refined' area

    Required params:
        accession(str): Genome accession to copy

    Returns:
        None
    """

    annotation = Path(f"data/refined/annotations/{accession}")
    fasta_genome = Path(f"data/refined/fasta/genomes/{accession}.fasta")
    fasta_proteins = Path(f"data/refined/fasta/proteins/{accession}.fasta")

    try:
        annotation.symlink_to(f"../../../data/full/annotations/{accession}")
    except FileExistsError as e:
        print(f"{e}: {accession} genome symlink exists")


    try:
        fasta_genome.symlink_to(f"../../../../data/full/fasta/genomes/{accession}.fasta")
    except FileExistsError as e:
        print(f"{e}: {accession} fasta genome symlink exists")

    try:
        fasta_proteins.symlink_to(f"../../../../data/full/fasta/proteins/{accession}.fasta")
    except FileExistsError as e:
        print(f"{e}: {accession} protein genome symlink exists")

def copy_target_proteins(accession):
    """
    Symlinks target protein sequences extracted from each genome for selected
    accession into 'refined' area 
    
    Required params:
        accession(str): genome accession to copy
    
    Returns:
        None
    """

    for gene in ["comP", "comQ", "comA", "comX", "rapP"]:
        source_file = Path(f"data/full/fasta/target_proteins/{gene}/{accession}.fasta")
        target_file = Path(f"data/refined/fasta/target_proteins/{gene}/{accession}.fasta")
        if source_file.exists():
            try:
                target_file.symlink_to(Path("../../../../../" / source_file))
            except FileExistsError as e:
                print(f"{e}: {accession} fasta protein symlink exists")
        else:
            print(f"{source_file} not found")

def refine_genomes(accession_file, bcgtree_config):
    """
    Creates subset of genomes which meet filtering criteria, and creates
    blast indexes of genomes and proteomes

    Required params:
        accessions_file: Path to file of accessions
        bcgtree_config: Path to write bcgtree config file
    
    Returns:
        None
    """

    data_root = Path("data/refined")

    annotations_dir = Path(data_root / "annotations")
    genome_fasta_dir =  Path(data_root / "fasta/genomes")
    protein_fasta_dir =  Path(data_root / "fasta/proteins")
    genome_blast_dir =  Path(data_root / "blast/genomes")
    protein_blast_dir =  Path(data_root / "blast/proteins")
    target_protein_dir = Path(data_root / "fasta/target_proteins")

    annotations_dir.mkdir(parents = True, exist_ok=True)
    genome_fasta_dir.mkdir(parents = True, exist_ok = True)
    protein_fasta_dir.mkdir(parents = True, exist_ok = True)
    genome_blast_dir.mkdir(parents = True, exist_ok = True)
    protein_blast_dir.mkdir(parents = True, exist_ok = True)
    target_protein_dir.mkdir(parents = True, exist_ok = True)

    for gene in ["comP", "comQ", "comA", "comX", "rapP"]:
        protein_dir = Path(target_protein_dir / gene)
        protein_dir.mkdir(exist_ok = True)

    df = pd.read_csv(accession_file)
    df['Accession'].apply(copy_genome)
    df['Accession'].apply(copy_target_proteins)

    blast_index(genome_fasta_dir, genome_blast_dir, SPECIES, NCBI_TAXID, 'nucl')
    blast_index(protein_fasta_dir, protein_blast_dir, SPECIES, NCBI_TAXID, 'prot')

    proteomes = [f"--proteome {x}={protein_fasta_dir}/{x}.fasta" for x in df['Accession']]
    proteomes = " ".join(proteomes)
    proteomes = "--proteome GCA_003868675=reference_sequences/outgroup/GCA_003868675.faa " + proteomes

    with open (bcgtree_config, 'w', encoding='UTF-8') as out_fh:
        out_fh.write(proteomes)
